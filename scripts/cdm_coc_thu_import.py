"""Import tọa độ cọc CDM thử từ Excel → JSON + SQLite (idempotent).

Nguồn: G:/My Drive/202605-TRUNG TAM HCM/CDM/TOA DO CDM-COC THU.xlsx
       Sheet "CDM THỬ" — 18 cọc (CỌC-01..CỌC-18)
       Cột: Point Name | Northing (m) | Easting (m) | Elevation (m) | Description

Output:
  - data/cdm_coc_thu_202605_TTHC.json
  - SQLite: bảng cdm_coc_thu (idempotent qua INSERT OR REPLACE)

Convention tọa độ: VN-2000 — northing_m, easting_m (giống cdm_toado).

Chạy: python scripts/cdm_coc_thu_import.py
"""
from __future__ import annotations

import json
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Thiếu openpyxl. Chạy: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

_ROOT  = Path(__file__).resolve().parent.parent
_XLSX  = Path(r"G:/My Drive/202605-TRUNG TAM HCM/CDM/TOA DO CDM-COC THU.xlsx")
_JSON  = _ROOT / "data" / "cdm_coc_thu_202605_TTHC.json"
_DB    = _ROOT / "data" / "TTHC.sqlite"
_SHEET = "CDM THỬ"


# ── Schema ───────────────────────────────────────────────────────────────────
def create_table(con: sqlite3.Connection) -> None:
    con.execute("""
        CREATE TABLE IF NOT EXISTS cdm_coc_thu (
            id          INTEGER PRIMARY KEY AUTOINCREMENT,
            point_name  INTEGER NOT NULL,
            code        TEXT    NOT NULL UNIQUE,
            northing_m  REAL    NOT NULL,
            easting_m   REAL    NOT NULL,
            elevation_m REAL    DEFAULT 0,
            description TEXT,
            source      TEXT    DEFAULT 'TOA DO CDM-COC THU.xlsx',
            created_at  TEXT    DEFAULT (datetime('now','localtime'))
        )
    """)
    con.execute("CREATE INDEX IF NOT EXISTS idx_cdm_coc_thu_code ON cdm_coc_thu(code)")


# ── Parser ───────────────────────────────────────────────────────────────────
def parse_xlsx(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if _SHEET not in wb.sheetnames:
        raise ValueError(f"Không tìm thấy sheet '{_SHEET}' trong {xlsx_path.name}")
    ws = wb[_SHEET]

    rows = []
    for i, raw in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1:
            continue  # header
        if raw[0] is None or raw[1] is None or raw[2] is None:
            continue
        pn = int(raw[0])
        nor = float(raw[1])
        eas = float(raw[2])
        elev = float(raw[3]) if raw[3] is not None else 0.0
        desc = str(raw[4]).strip() if raw[4] is not None else f"CỌC-{pn:02d}"
        # code = CỌC-NN
        code = desc if desc.startswith("CỌC") else f"CỌC-{pn:02d}"
        rows.append({
            "point_name":  pn,
            "code":        code,
            "northing_m":  round(nor, 3),
            "easting_m":   round(eas, 3),
            "elevation_m": round(elev, 3),
            "description": desc,
        })
    wb.close()
    return rows


# ── Save JSON ────────────────────────────────────────────────────────────────
def save_json(rows: list[dict], xlsx_path: Path) -> None:
    _JSON.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "_meta": {
            "source":   xlsx_path.name,
            "sheet":    _SHEET,
            "updated":  datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "n_piles":  len(rows),
            "crs":      "VN-2000",
            "note":     "Cọc CDM thử nghiệm — 18 cọc, cao độ mặt = 0 (chưa khảo sát chi tiết)",
        },
        "piles": rows,
    }
    _JSON.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


# ── Save SQLite ──────────────────────────────────────────────────────────────
def save_sqlite(rows: list[dict], db_path: Path = _DB) -> None:
    con = sqlite3.connect(db_path)
    try:
        create_table(con)
        for r in rows:
            con.execute(
                """INSERT INTO cdm_coc_thu
                   (point_name, code, northing_m, easting_m, elevation_m, description)
                   VALUES (?, ?, ?, ?, ?, ?)
                   ON CONFLICT(code) DO UPDATE SET
                     point_name  = excluded.point_name,
                     northing_m  = excluded.northing_m,
                     easting_m   = excluded.easting_m,
                     elevation_m = excluded.elevation_m,
                     description = excluded.description""",
                (r["point_name"], r["code"], r["northing_m"],
                 r["easting_m"], r["elevation_m"], r["description"]),
            )
        con.commit()
    finally:
        con.close()


# ── Main ─────────────────────────────────────────────────────────────────────
def main() -> None:
    if not _XLSX.exists():
        print(f"[ERROR] Không tìm thấy file: {_XLSX}", file=sys.stderr)
        sys.exit(2)

    print(f"[1/3] Đọc {_XLSX.name} sheet '{_SHEET}'...")
    rows = parse_xlsx(_XLSX)
    print(f"      → {len(rows)} cọc thử")

    print(f"[2/3] Lưu JSON: {_JSON.relative_to(_ROOT)}")
    save_json(rows, _XLSX)

    print(f"[3/3] Cập nhật SQLite: {_DB.relative_to(_ROOT)}")
    save_sqlite(rows)

    # Tóm tắt
    print("\n=== TÓM TẮT ===")
    n = sum(1 for _ in rows)
    print(f"Số cọc thử: {n}")
    print(f"Phạm vi Northing: {min(r['northing_m'] for r in rows):,.1f} – "
          f"{max(r['northing_m'] for r in rows):,.1f} m")
    print(f"Phạm vi Easting:  {min(r['easting_m'] for r in rows):,.1f} – "
          f"{max(r['easting_m'] for r in rows):,.1f} m")
    print("\nHoàn tất.")


if __name__ == "__main__":
    main()
