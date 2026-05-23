"""cdm_import_excel.py
Import dữ liệu thi công + thí nghiệm CDM từ file Excel đã điền vào SQLite.

Chạy:
    python -X utf8 scripts/cdm_import_excel.py [path_excel]
    (không truyền path → tự tìm file CDM_MauBang_*.xlsx mới nhất trong data/)
Idempotent — UPDATE cdm_thi_cong, INSERT OR REPLACE cdm_kiem_tra.
"""
import sqlite3, sys, re
from pathlib import Path
from datetime import date, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Cần cài openpyxl: pip install openpyxl")

ROOT = Path(__file__).parent.parent
DB   = ROOT / "data" / "TTHC.sqlite"

# ── Column positions (1-based) cho sheet "Thi công CDM" ───────────────────────
# A(1)=id  B=zone  C=point_name  D=northing  E=easting  F=bh
# G=L_design  H=z_top_design  I=z_tip_design  J=D  K=xi_mang  L=nc  ← design (skip)
# M(13)=status  N=ngay  O=to_doi  P=may  Q=z_top_actual  R=z_tip_actual
# S=L_actual  T=luong_xi_mang_kg  U=ap_luc  V=toc_do_khoan  W=toc_do_rut  X=ghi_chu
TC_ID_COL    = 1   # cột A = id (khóa tra cứu)
TC_ZONE_COL  = 2
TC_INPUT_FIELDS = [
    (13, "status"),
    (14, "ngay_thi_cong"),
    (15, "to_doi"),
    (16, "may_thi_cong"),
    (17, "z_top_actual_m"),
    (18, "z_tip_actual_m"),
    (19, "L_actual_m"),
    (20, "luong_xi_mang_kg"),
    (21, "ap_luc_bom_bar"),
    (22, "toc_do_khoan_m_ph"),
    (23, "toc_do_rut_m_ph"),
    (24, "ghi_chu"),
]
TC_STATUS_VALID = {"chua_thi_cong","dang_thi_cong","hoan_thanh","khong_dat","tam_dung"}

# ── Column positions cho sheet "Thí nghiệm CDM" ───────────────────────────────
# A=zone  B=point_name  C=loai_tn  D=phuong_phap  E=ngay_lay  F=ngay_tn
# G=tuoi_ngay  H=do_sau  I=z_sample  J=qu  K=qu_yc  L=dat_yeu_cau
# M=ham_luong_xm  N=nc  O=ghi_chu
KT_FIELDS = [
    (1,  "zone"),
    (2,  "point_name"),
    (3,  "loai_tn"),
    (4,  "phuong_phap_lay_mau"),
    (5,  "ngay_lay_mau"),
    (6,  "ngay_thu_nghiem"),
    (7,  "tuoi_ngay"),
    (8,  "do_sau_tu_dinh_m"),
    (9,  "z_sample_m"),
    (10, "qu_kPa"),
    (11, "qu_yc_kPa"),
    (12, "dat_yeu_cau"),
    (13, "ham_luong_xi_mang_pct"),
    (14, "ty_le_nc"),
    (15, "ghi_chu"),
]

# ── Helpers ────────────────────────────────────────────────────────────────────
def _clean(v):
    """Normalize cell value: strip string, None for blank."""
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (date, datetime)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip() or None

def _parse_dat(v):
    """Đạt / Không đạt / None → 1 / 0 / None"""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("1", "Đạt", "Dat"):
        return 1
    if s in ("0", "Không đạt", "Khong dat"):
        return 0
    return None

def _find_excel(path_arg=None) -> Path:
    if path_arg:
        p = Path(path_arg)
        if not p.exists():
            sys.exit(f"Không tìm thấy file: {path_arg}")
        return p
    candidates = sorted((ROOT / "data").glob("CDM_MauBang_*.xlsx"), reverse=True)
    if not candidates:
        sys.exit("Không tìm thấy CDM_MauBang_*.xlsx trong data/. Truyền path thủ công.")
    return candidates[0]

# ── Import sheet Thi công CDM ──────────────────────────────────────────────────
def import_sheet_tc(ws, con) -> tuple[int, int]:
    """Returns (n_updated, n_skipped)"""
    updated = skipped = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row in ws.iter_rows(min_row=2, values_only=True):
        rec_id = _clean(row[TC_ID_COL - 1])
        if rec_id is None:
            continue
        try:
            rec_id = int(rec_id)
        except (ValueError, TypeError):
            continue

        # Chỉ update các cột nhập tay (M..X)
        updates: dict[str, object] = {}
        for col_idx, field in TC_INPUT_FIELDS:
            raw = _clean(row[col_idx - 1])
            if field == "status":
                if raw and raw not in TC_STATUS_VALID:
                    raw = None   # bỏ qua giá trị không hợp lệ
            updates[field] = raw

        # Bỏ qua nếu không có gì nhập (tất cả None)
        non_null = {k: v for k, v in updates.items() if v is not None}
        if not non_null:
            skipped += 1
            continue

        set_parts = ", ".join(f"{k}=?" for k in non_null) + ", updated_at=?"
        vals = list(non_null.values()) + [now, rec_id]
        con.execute(f"UPDATE cdm_thi_cong SET {set_parts} WHERE id=?", vals)
        updated += 1

    con.commit()
    return updated, skipped

# ── Import sheet Thí nghiệm CDM ───────────────────────────────────────────────
def import_sheet_kt(ws, con) -> tuple[int, int]:
    """Returns (n_inserted, n_skipped)"""
    # Xây lookup point_name → thi_cong_id
    tc_map = {(r[0], r[1]): r[2] for r in con.execute(
        "SELECT zone, point_name, id FROM cdm_thi_cong"
    ).fetchall()}

    inserted = skipped = 0
    rows_to_insert = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        zone  = _clean(row[0])
        pname = _clean(row[1])
        if not zone or not pname:
            continue

        # Tìm thi_cong_id
        tc_id = tc_map.get((zone, pname))

        record: dict[str, object] = {"thi_cong_id": tc_id}
        for col_idx, field in KT_FIELDS:
            raw = _clean(row[col_idx - 1])
            if field == "dat_yeu_cau":
                raw = _parse_dat(raw)
            record[field] = raw

        # qu_kPa bắt buộc (cột kết quả chính)
        if record.get("qu_kPa") is None and record.get("loai_tn") is None:
            skipped += 1
            continue

        rows_to_insert.append(record)
        inserted += 1

    fields_list = ["thi_cong_id"] + [f for _, f in KT_FIELDS]
    ph = ",".join("?" * len(fields_list))
    con.executemany(
        f"INSERT INTO cdm_kiem_tra ({','.join(fields_list)}) VALUES ({ph})",
        [[r.get(f) for f in fields_list] for r in rows_to_insert],
    )
    con.commit()
    return inserted, skipped

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    xlsx_path = _find_excel(sys.argv[1] if len(sys.argv) > 1 else None)
    print(f"Đọc file: {xlsx_path.name}")

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    sheet_names = wb.sheetnames
    print(f"  Sheets: {sheet_names}")

    con = sqlite3.connect(str(DB))
    con.execute("PRAGMA journal_mode=WAL")

    # Sheet 1: Thi công CDM
    if "Thi công CDM" in sheet_names:
        ws_tc = wb["Thi công CDM"]
        n_upd, n_skip = import_sheet_tc(ws_tc, con)
        print(f"\nSheet 'Thi công CDM':")
        print(f"  Đã cập nhật: {n_upd:,} cọc")
        print(f"  Bỏ qua (không có dữ liệu nhập): {n_skip:,} cọc")
    else:
        print("  Không tìm thấy sheet 'Thi công CDM'")

    # Sheet 2: Thí nghiệm CDM
    if "Thí nghiệm CDM" in sheet_names:
        ws_kt = wb["Thí nghiệm CDM"]
        n_ins, n_skip2 = import_sheet_kt(ws_kt, con)
        print(f"\nSheet 'Thí nghiệm CDM':")
        print(f"  Đã chèn: {n_ins:,} mẫu")
        print(f"  Bỏ qua (thiếu qu/loại TN): {n_skip2:,} dòng")
    else:
        print("  Không tìm thấy sheet 'Thí nghiệm CDM'")

    # Thống kê sau import
    print("\nPhân bố trạng thái sau import:")
    for st, zone, n in con.execute(
        "SELECT status, zone, COUNT(*) FROM cdm_thi_cong "
        "GROUP BY status, zone ORDER BY status, zone"
    ).fetchall():
        print(f"  {zone:12s}  {st:20s}  {n:,}")

    n_kt = con.execute("SELECT COUNT(*) FROM cdm_kiem_tra").fetchone()[0]
    print(f"\nTổng mẫu kiểm tra cdm_kiem_tra: {n_kt:,}")

    con.close()
    print("\nHoàn thành.")

if __name__ == "__main__":
    main()
