"""Tạo file Excel tính móng trụ CDM per hố khoan từ template LK1 + dữ liệu SQLite.

Giữ NGUYÊN công thức + định dạng template "TINH MONG TRU CDM - LK1.xlsx", chỉ thay
ô ĐẦU VÀO:
  - C3: tên hố khoan
  - Cấu tạo tải đắp (hàng 8-12): theo cấu tạo TTHC (Σ=1.9m → q=40.8 kPa)
  - Hình học: O25 CDTK, O26 CDTN, I25 CD1, I26 CD2(mũi), O27 CDNN, V35 (neo cao độ)
  - Bảng địa chất per-1m (hàng 33-62): D,E,F,I,K,M,N,O,R + công thức G,H,J,P,Q

Nguồn dữ liệu: data/TTHC.sqlite (boreholes, layers, lab_tests, vane_shear_tests,
cdm_zone_design_results cho tip_depth).

Dùng: python scripts/cdm_excel_per_hk.py KE-HK1
      python scripts/cdm_excel_per_hk.py --all
"""
from __future__ import annotations

import sqlite3
import sys
from pathlib import Path

from openpyxl import load_workbook

_ROOT = Path(__file__).resolve().parent.parent
_DB = _ROOT / "data" / "TTHC.sqlite"
_TEMPLATE = Path(r"G:\My Drive\202605-TRUNG TAM HCM\CDM\TINH MONG TRU CDM - LK1.xlsx")
_OUTDIR = Path(r"G:\My Drive\202605-TRUNG TAM HCM\CDM")

ROW0 = 33           # hàng địa chất đầu
ROW_MAX = 62        # hàng địa chất cuối (downstream refs hỗ trợ tới đây)
NDEPTH = ROW_MAX - ROW0 + 1   # = 30 hàng (30 m)

CDTK = 2.7          # cao độ thiết kế (tvtk_cdm_config.settlement_design_elev_m)
CD1 = 0.8           # cao độ đỉnh trụ (top_elev_m)
GWL_ELEV = 0.0      # cao độ mực nước ngầm
QUCK = 800.0        # cường độ trụ thiết kế
# Cấu tạo tải đắp TTHC: (label, gamma, thickness)
FILL = [("He cát đắp", 18.0, 0.7), ("Áo đường", 24.0, 0.8),
        ("Hse đệm cát", 22.5, 0.4)]

SAND_SYMBOLS = {"F", "2a", "2b", "2c", "3a", "3b", "3c", "4", "5", "5a",
                "5b", "6", "7", "8"}


def _con():
    c = sqlite3.connect(str(_DB)); c.row_factory = sqlite3.Row
    return c


def _bh_natural(con, bh):
    r = con.execute("SELECT elevation_m FROM boreholes WHERE name=?", (bh,)).fetchone()
    return float(r["elevation_m"]) if r and r["elevation_m"] is not None else 0.0


def _tip_depth(con, bh):
    r = con.execute(
        "SELECT tip_depth_m FROM cdm_zone_design_results "
        "WHERE bh_name=? AND delta_S_cm=30 AND zone_code LIKE 'KE%' LIMIT 1", (bh,)
    ).fetchone()
    if r and r["tip_depth_m"] is not None:
        return float(r["tip_depth_m"])
    # fallback: H_soft + 1 m ngàm
    r2 = con.execute(
        "SELECT SUM(depth_bot_m-depth_top_m) s FROM layers l JOIN boreholes b "
        "ON l.borehole_id=b.id WHERE b.name=? AND l.symbol IN ('1','1b','2','XMD')",
        (bh,)).fetchone()
    return (float(r2["s"]) + 1.0) if r2 and r2["s"] else 20.0


def _layer_rows(con, bh):
    rows = con.execute("""
        SELECT l.symbol, l.depth_top_m, l.depth_bot_m,
               ROUND(AVG(lt.gamma_kNm3),2) g, ROUND(AVG(lt.e0),3) e0,
               ROUND(AVG(lt.Cc),4) Cc, ROUND(AVG(lt.Cs),4) Cs,
               ROUND(AVG(lt.PC_kPa),1) PC, ROUND(AVG(lt.Cu_UU_kPa),1) Cuu
        FROM layers l JOIN boreholes b ON l.borehole_id=b.id
        LEFT JOIN lab_tests lt ON lt.borehole_id=b.id
            AND lt.depth_from_m>=l.depth_top_m AND lt.depth_from_m<l.depth_bot_m
        WHERE b.name=? GROUP BY l.id ORDER BY l.depth_top_m
    """, (bh,)).fetchall()
    return [dict(r) for r in rows]


def _vst(con, bh):
    rows = con.execute("""
        SELECT v.depth_m d, v.Su_kPa su FROM vane_shear_tests v
        JOIN vst_locations loc ON v.vst_loc_id=loc.id
        WHERE loc.name=? AND v.Su_kPa>0 ORDER BY v.depth_m
    """, (bh,)).fetchall()
    return [(float(r["d"]), float(r["su"])) for r in rows]


def _interp(pts, x):
    if not pts:
        return None
    if x <= pts[0][0]:
        return pts[0][1]
    if x >= pts[-1][0]:
        return pts[-1][1]
    for (x0, y0), (x1, y1) in zip(pts, pts[1:]):
        if x0 <= x <= x1:
            return y0 + (y1 - y0) * (x - x0) / (x1 - x0) if x1 > x0 else y0
    return pts[-1][1]


def _nearest_gamma(layers):
    for l in layers:
        if l.get("g"):
            return l["g"]
    return 17.0


def build_profile(con, bh):
    """List per-1m dict: {D,E,gamma,Su,e0,Cc,Cr,PC} cho NDEPTH hàng."""
    layers = _layer_rows(con, bh)
    vst = _vst(con, bh)
    g_default = _nearest_gamma(layers)
    out = []
    for i in range(NDEPTH):
        mid = i + 0.5
        lay = next((l for l in layers
                    if l["depth_top_m"] <= mid < l["depth_bot_m"]), None)
        if lay is None:
            lay = layers[-1] if layers else {}
        sym = (lay.get("symbol") or "").strip()
        is_sand = sym in SAND_SYMBOLS
        lid = (layers.index(lay) + 1) if lay in layers else 1
        gamma = lay.get("g") or g_default
        if is_sand:
            out.append({"D": lid, "E": "Cát", "gamma": round(gamma, 2),
                        "Su": 0.0, "e0": lay.get("e0") or 0.6,
                        "Cc": 0.0, "Cr": 0.0, "PC": 0.0})
        else:
            su = _interp(vst, mid)
            if su is None:
                su = lay.get("Cuu") or 10.0
            out.append({"D": lid, "E": "Sét", "gamma": round(gamma, 2),
                        "Su": round(su, 1), "e0": lay.get("e0") or 1.4,
                        "Cc": lay.get("Cc") or 0.0, "Cr": lay.get("Cs") or 0.0,
                        "PC": lay.get("PC") or 0.0})
    return out


def fill_one(bh: str, out_path: Path) -> dict:
    con = _con()
    try:
        natural = _bh_natural(con, bh)
        tip = _tip_depth(con, bh)
        prof = build_profile(con, bh)
    finally:
        con.close()

    n = int(bh.replace("KE-HK", ""))
    suffix = f"HK{n:02d}"

    wb = load_workbook(str(_TEMPLATE), data_only=False)
    ws = wb["(1)"]

    # Tên HK
    ws["C3"] = f"Áp dụng cho hố khoan {suffix}"

    # Cấu tạo tải đắp (hàng 8-12) → TTHC; phần dư đặt 0
    for k, (label, g, h) in enumerate(FILL):
        r = 8 + k
        ws.cell(r, 5).value = label      # E
        ws.cell(r, 9).value = g          # I gamma
        ws.cell(r, 10).value = h         # J thickness
    for r in range(8 + len(FILL), 13):   # các hàng còn lại = 0
        ws.cell(r, 10).value = 0

    # Hình học
    ws["O25"] = CDTK
    ws["O26"] = round(natural, 2)        # CDTN (override công thức)
    ws["I25"] = CD1                      # CD1 đỉnh trụ (override)
    ws["I26"] = round(natural - tip, 2)  # CD2 đáy trụ (mũi)
    ws["O27"] = GWL_ELEV                 # CDNN
    ws["V35"] = round(natural - 1.0, 2)  # neo chuỗi cao độ địa chất
    ws["I23"] = 0.8                      # D
    ws["I24"] = 1.8                      # S
    ws["J65"] = 1.8                      # S (khối trụ)
    ws["L65"] = 1                        # lưới vuông
    ws["J69"] = QUCK                     # quck

    # Bảng địa chất per-1m (hàng 33..ROW_MAX)
    for i in range(NDEPTH):
        r = ROW0 + i
        p = prof[i]
        ws.cell(r, 4).value = p["D"]                 # D lớp
        ws.cell(r, 5).value = p["E"]                 # E loại đất
        ws.cell(r, 6).value = 1.0                    # F bề dày
        ws.cell(r, 7).value = ("=+F33" if r == ROW0 else f"=+G{r-1}+F{r}")   # G
        ws.cell(r, 8).value = ("=+$V$35" if r == ROW0 else f"=+H{r-1}-F{r}")  # H
        ws.cell(r, 9).value = p["gamma"]             # I gamma
        ws.cell(r, 10).value = f"=+I{r}-10"          # J gamma_dn
        ws.cell(r, 11).value = p["Su"]               # K Su
        ws.cell(r, 13).value = p["e0"]               # M e0
        ws.cell(r, 14).value = p["Cc"]               # N Cc
        ws.cell(r, 15).value = p["Cr"]               # O Cr
        ws.cell(r, 16).value = f"=K{r}*200"          # P E
        ws.cell(r, 17).value = f"=R{r}*10"           # Q spz
        ws.cell(r, 18).value = round(p["PC"] / 10.0, 4)  # R = PC/10

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return {"bh": bh, "suffix": suffix, "natural": natural, "tip": tip,
            "file": str(out_path), "n_layers": NDEPTH}


def main():
    args = sys.argv[1:]
    if not args:
        args = ["KE-HK1"]
    if args == ["--all"]:
        bhs = [f"KE-HK{i}" for i in range(1, 13)]
    else:
        bhs = args
    for bh in bhs:
        n = int(bh.replace("KE-HK", ""))
        out = _OUTDIR / f"TINH MONG TRU CDM - HK{n:02d}.xlsx"
        info = fill_one(bh, out)
        print(f"[OK] {bh} -> {info['file']}  (natural={info['natural']:.2f} "
              f"tip={info['tip']:.1f})")


if __name__ == "__main__":
    main()
