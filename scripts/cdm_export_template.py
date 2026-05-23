"""cdm_export_template.py
Xuất mẫu bảng thi công + thí nghiệm CDM ra Excel.
Sheet 1: cdm_thi_cong  (toàn bộ cọc, kèm cột nhập tay)
Sheet 2: cdm_kiem_tra  (bảng nhập kết quả TN)
Sheet 3: Hướng dẫn
Chạy: python -X utf8 scripts/cdm_export_template.py
"""
import sqlite3
from pathlib import Path
from datetime import date

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    import sys; sys.exit("Cài openpyxl: pip install openpyxl")

ROOT = Path(__file__).parent.parent
DB   = ROOT / "data" / "TTHC.sqlite"
OUT  = ROOT / "data" / f"CDM_MauBang_{date.today().isoformat()}.xlsx"

# ── Màu sắc ──────────────────────────────────────────────────────────────────
CLR = {
    "hdr_tk":    "1565C0",   # xanh đậm — cột thiết kế (read-only)
    "hdr_nhap":  "2E7D32",   # xanh lá — cột nhập tay
    "hdr_tn":    "6A1B9A",   # tím — bảng TN
    "hdr_kq":    "BF360C",   # cam đậm — kết quả TN
    "row_ke":    "E3F2FD",   # nền xanh nhạt — khu Kè
    "row_cv":    "F3E5F5",   # nền tím nhạt — Công viên
    "row_tn":    "F3E5F5",
    "white":     "FFFFFF",
    "yellow":    "FFFDE7",
    "lock":      "ECEFF1",   # xám nhạt — cột chỉ đọc
}

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Calibri")

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def _border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

# ── Sheet 1: cdm_thi_cong ─────────────────────────────────────────────────────
COLS_TC = [
    # (header, db_field, width, color_group, number_format)
    ("STT",                    "id",               5,  "tk",  "0"),
    ("Khu vực",                "zone",             10, "tk",  "@"),
    ("Số hiệu cọc",            "point_name",       10, "tk",  "@"),
    ("Northing (m)",           "northing_m",       14, "tk",  "0.000"),
    ("Easting (m)",            "easting_m",        14, "tk",  "0.000"),
    ("HK điều hành",           "bh_dieu_hanh",     14, "tk",  "@"),
    ("L thiết kế (m)",         "L_design_m",       12, "tk",  "0.00"),
    ("Cao độ đỉnh TK (m)",     "z_top_design_m",   14, "tk",  "0.00"),
    ("Cao độ mũi TK (m)",      "z_tip_design_m",   14, "tk",  "0.00"),
    ("D (m)",                  "D_m",              8,  "tk",  "0.00"),
    ("Xi măng TK (kg/m)",      "ham_luong_xi_mang",14, "tk",  "0"),
    ("w/c TK",                 "ty_le_nc",         8,  "tk",  "0.00"),
    # Nhập tay
    ("Trạng thái",             "status",           16, "nhap","@"),
    ("Ngày thi công",          "ngay_thi_cong",    14, "nhap","YYYY-MM-DD"),
    ("Tổ đội",                 "to_doi",           14, "nhap","@"),
    ("Máy thi công",           "may_thi_cong",     14, "nhap","@"),
    ("Cao độ đỉnh thực (m)",   "z_top_actual_m",   14, "nhap","0.00"),
    ("Cao độ mũi thực (m)",    "z_tip_actual_m",   14, "nhap","0.00"),
    ("L thực tế (m)",          "L_actual_m",       12, "nhap","0.00"),
    ("Xi măng thực tế (kg)",   "luong_xi_mang_kg", 16, "nhap","0"),
    ("Áp lực bơm (bar)",       "ap_luc_bom_bar",   14, "nhap","0.0"),
    ("Tốc độ khoan (m/ph)",    "toc_do_khoan_m_ph",16, "nhap","0.00"),
    ("Tốc độ rút (m/ph)",      "toc_do_rut_m_ph",  14, "nhap","0.00"),
    ("Ghi chú",                "ghi_chu",          20, "nhap","@"),
]

COLS_KT = [
    ("Khu vực",                "zone",                   10, "tn",  "@"),
    ("Số hiệu cọc",            "point_name",             10, "tn",  "@"),
    ("Loại TN",                "loai_tn",                14, "tn",  "@"),
    ("Phương pháp lấy mẫu",    "phuong_phap_lay_mau",    18, "tn",  "@"),
    ("Ngày lấy mẫu",           "ngay_lay_mau",           14, "tn",  "YYYY-MM-DD"),
    ("Ngày thử nghiệm",        "ngay_thu_nghiem",        16, "tn",  "YYYY-MM-DD"),
    ("Tuổi mẫu (ngày)",        "tuoi_ngay",              14, "tn",  "0"),
    ("Độ sâu từ đỉnh (m)",     "do_sau_tu_dinh_m",       16, "tn",  "0.00"),
    ("Cao độ mẫu (m)",         "z_sample_m",             14, "tn",  "0.00"),
    ("qu thực tế (kPa)",       "qu_kPa",                 14, "kq",  "0.0"),
    ("qu yêu cầu (kPa)",       "qu_yc_kPa",              14, "kq",  "0.0"),
    ("Đạt/Không đạt",          "dat_yeu_cau",            14, "kq",  "@"),
    ("Hàm lượng xi măng (%)",  "ham_luong_xi_mang_pct",  18, "tn",  "0.0"),
    ("w/c thực tế",            "ty_le_nc",               10, "tn",  "0.00"),
    ("Ghi chú",                "ghi_chu",                20, "tn",  "@"),
]

def _write_header(ws, cols, row=1):
    for ci, (hdr, fld, width, grp, fmt) in enumerate(cols, 1):
        cl = get_column_letter(ci)
        cell = ws.cell(row=row, column=ci, value=hdr)
        color = CLR[f"hdr_{grp}"]
        cell.fill     = _fill(color)
        cell.font     = _font(bold=True, color="FFFFFF", size=10)
        cell.alignment = _center()
        cell.border   = _border()
        ws.column_dimensions[cl].width = width
    ws.row_dimensions[row].height = 36

def _write_row(ws, row_idx, values, cols, zone="KE"):
    bg = CLR["row_ke"] if zone == "KE" else CLR["row_cv"]
    for ci, (val, (hdr, fld, width, grp, fmt)) in enumerate(
            zip(values, cols), 1):
        cell = ws.cell(row=row_idx, column=ci, value=val)
        cell.border = _border()
        cell.alignment = _left()
        if grp == "tk":
            cell.fill = _fill(CLR["lock"])
            cell.font = _font(color="424242", size=9)
        else:
            cell.fill = _fill(bg)
            cell.font = _font(size=9)
        if fmt not in ("@", "YYYY-MM-DD") and val is not None:
            cell.number_format = fmt

def build_sheet_tc(wb, con):
    ws = wb.create_sheet("Thi công CDM")
    ws.freeze_panes = "M2"      # freeze đến cột nhập đầu tiên

    # Header
    _write_header(ws, COLS_TC)

    # Data
    fields = [c[1] for c in COLS_TC]
    sel_fields = ", ".join(
        f"id" if f == "id" else f for f in fields
    )
    rows = con.execute(
        f"SELECT {sel_fields} FROM cdm_thi_cong ORDER BY zone, CAST(point_name AS INTEGER)"
    ).fetchall()

    for ri, row in enumerate(rows, 2):
        zone = row[1] if row[1] else "KE"
        _write_row(ws, ri, row, COLS_TC, zone)

    # Data validation: status dropdown
    status_opts = '"chua_thi_cong,dang_thi_cong,hoan_thanh,khong_dat,tam_dung"'
    dv_status = DataValidation(
        type="list", formula1=status_opts, allow_blank=True,
        showErrorMessage=True,
        errorTitle="Giá trị không hợp lệ",
        error="Chọn từ danh sách: chua_thi_cong / dang_thi_cong / hoan_thanh / khong_dat / tam_dung",
    )
    ws.add_data_validation(dv_status)
    col_status = get_column_letter(fields.index("status") + 1)
    dv_status.sqref = f"{col_status}2:{col_status}{len(rows)+1}"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS_TC))}1"
    ws.sheet_view.showGridLines = True
    return ws

def build_sheet_kt(wb):
    ws = wb.create_sheet("Thí nghiệm CDM")
    _write_header(ws, COLS_KT)

    # Mẫu 5 dòng ví dụ
    samples = [
        ("KE",       "1025", "UCS",  "wet_grab",   "2026-07-01","2026-07-29", 28, 5.0, -18.5, 800, 500, "Đạt",  12, 0.55, ""),
        ("KE",       "1025", "UCS",  "wet_grab",   "2026-07-01","2026-07-29", 28,10.0, -23.5, 750, 500, "Đạt",  12, 0.55, ""),
        ("KE",       "2310", "UCS",  "coring",     "2026-07-05","2026-08-02", 28, 3.0, -16.0, 480, 500, "Không đạt", 10, 0.60, "Mẫu bị nứt"),
        ("CONG_VIEN","4700", "UCS",  "wet_grab",   "2026-07-10","2026-08-07", 28, 6.0, -14.5, 610, 500, "Đạt",  12, 0.55, ""),
        ("CONG_VIEN","8820", "core", "khoan_lay_mau","2026-07-12","2026-08-09",28, 4.0, -13.2, 530, 500, "Đạt",  12, 0.55, ""),
    ]
    for ri, row in enumerate(samples, 2):
        zone = row[0]
        bg = CLR["row_ke"] if zone == "KE" else CLR["row_cv"]
        for ci, (val, (hdr, fld, width, grp, fmt)) in enumerate(
                zip(row, COLS_KT), 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = _border()
            cell.alignment = _left()
            cell.fill = _fill(bg if grp == "tn" else CLR["yellow"])
            cell.font = _font(size=9, color="424242")

    # Validation: loai_tn
    dv_loai = DataValidation(
        type="list",
        formula1='"UCS,core,wet_grab,field_vane"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_loai)
    dv_loai.sqref = "C2:C5000"

    # Validation: dat_yeu_cau
    dv_dat = DataValidation(
        type="list",
        formula1='"Đạt,Không đạt,Chưa đánh giá"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_dat)
    dv_dat.sqref = "L2:L5000"

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS_KT))}1"
    ws.freeze_panes = "A2"
    return ws

def build_sheet_huongdan(wb):
    ws = wb.create_sheet("Hướng dẫn")
    ws.sheet_view.showGridLines = False

    def _h(row, text, bold=True, size=12, color="1565C0"):
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=color, name="Calibri")
        cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 22 if bold else 18
        return cell

    def _row(row, col, val, bold=False):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(size=10, bold=bold, name="Calibri")
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 55

    _h(1, "HƯỚNG DẪN SỬ DỤNG BẢNG THEO DÕI THI CÔNG CDM", size=13)
    _h(3, "Sheet 1 — Thi công CDM", size=11)
    items_tc = [
        ("Cột xanh đậm (màu xám nhạt)","Thông tin thiết kế — đọc từ hệ thống, KHÔNG sửa"),
        ("Cột xanh lá","Nhập tay sau khi thi công"),
        ("Trạng thái","chua_thi_cong → dang_thi_cong → hoan_thanh / khong_dat"),
        ("Ngày thi công","Định dạng YYYY-MM-DD (ví dụ: 2026-07-15)"),
        ("Cao độ mũi thực (m)","Âm = dưới mực 0.00m (ví dụ: −22.5)"),
        ("L thực tế (m)","= Cao độ đỉnh thực − Cao độ mũi thực"),
        ("Xi măng thực tế (kg)","Tổng lượng xi măng đã bơm cho 1 cọc"),
    ]
    for ri, (k, v) in enumerate(items_tc, 4):
        _row(ri, 1, k, bold=True)
        _row(ri, 2, v)

    _h(12, "Sheet 2 — Thí nghiệm CDM", size=11)
    items_kt = [
        ("Loại TN","UCS = nén nở hông; core = khoan lõi; wet_grab = lấy mẫu ướt"),
        ("Tuổi mẫu (ngày)","Thường: 7, 14, 28, 60, 90 ngày"),
        ("qu thực tế (kPa)","Kết quả thí nghiệm nén đơn trục"),
        ("qu yêu cầu (kPa)","Theo thiết kế (thường 500 kPa ở 28 ngày)"),
        ("Đạt/Không đạt","Chọn từ danh sách dropdown"),
        ("Hàm lượng xi măng (%)","% xi măng theo khối lượng đất khô"),
    ]
    for ri, (k, v) in enumerate(items_kt, 13):
        _row(ri, 1, k, bold=True)
        _row(ri, 2, v)

    _h(20, "Quy trình nhập dữ liệu", size=11)
    steps = [
        "1. Lọc cột 'Khu vực' hoặc 'HK điều hành' để chọn khu vực làm việc.",
        "2. Cập nhật 'Trạng thái' → 'dang_thi_cong' khi bắt đầu thi công.",
        "3. Sau khi xong, nhập đầy đủ các cột màu xanh lá → đổi status = 'hoan_thanh'.",
        "4. Sheet TN: nhập kết quả qu sau khi có báo cáo phòng thí nghiệm.",
        "5. Import lại vào hệ thống bằng script cdm_import_excel.py (sắp có).",
    ]
    for ri, s in enumerate(steps, 21):
        _row(ri, 1, s)
        ws.row_dimensions[ri].height = 18

    return ws

# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    con = sqlite3.connect(str(DB))
    wb  = openpyxl.Workbook()
    wb.remove(wb.active)          # xóa sheet mặc định

    print("Xuất sheet Thi công CDM ...")
    ws_tc = build_sheet_tc(wb, con)
    n_tc = con.execute("SELECT COUNT(*) FROM cdm_thi_cong").fetchone()[0]

    print("Xuất sheet Thí nghiệm CDM ...")
    ws_kt = build_sheet_kt(wb)

    print("Xuất sheet Hướng dẫn ...")
    build_sheet_huongdan(wb)

    con.close()

    # Tab order
    ws_tc.sheet_state    = "visible"
    ws_kt.sheet_state    = "visible"

    wb.save(str(OUT))
    print(f"\nFile: {OUT.name}")
    print(f"  Sheet 'Thi công CDM':   {n_tc:,} cọc")
    print(f"  Sheet 'Thí nghiệm CDM': 5 dòng mẫu")
    print(f"  Sheet 'Hướng dẫn':      hướng dẫn sử dụng")
    print("Xong.")

if __name__ == "__main__":
    main()
