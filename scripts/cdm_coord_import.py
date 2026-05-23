"""cdm_coord_import.py
Import tọa độ cọc CDM từ Excel 'TOA DO CDM.xlsx' → SQLite cdm_toado + JSON.
Chạy: python -X utf8 scripts/cdm_coord_import.py
"""
import json, sqlite3, sys
from pathlib import Path
try:
    import openpyxl
except ImportError:
    sys.exit("Cài openpyxl trước: pip install openpyxl")

ROOT     = Path(__file__).parent.parent
DB       = ROOT / "data" / "TTHC.sqlite"
XLSX     = Path(r"G:\My Drive\202605-TRUNG TAM HCM\CDM\TOA DO CDM.xlsx")
JSON_OUT = ROOT / "data" / "cdm_toado_202605_TTHC.json"

# Tên sheet → zone code
SHEET_ZONE = {
    "CONG VIEN": "CONG_VIEN",
    "KE":        "KE",
}

# Lọc tọa độ hợp lệ VN-2000 (khu vực TP.HCM)
def _valid(n, e):
    return (n is not None and e is not None
            and 1_100_000 < float(n) < 1_300_000
            and 500_000 < float(e) < 700_000)

# ── Parse Excel ────────────────────────────────────────────────────────────────
def parse_xlsx(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    records = []
    for sheet_name, zone in SHEET_ZONE.items():
        if sheet_name not in wb.sheetnames:
            print(f"  [skip] sheet '{sheet_name}' không tồn tại")
            continue
        ws = wb[sheet_name]
        n_ok = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 3:
                continue
            pt_name, northing, easting = row[0], row[1], row[2]
            if not _valid(northing, easting):
                continue
            records.append({
                "zone":       zone,
                "point_name": str(pt_name) if pt_name is not None else None,
                "northing_m": round(float(northing), 3),
                "easting_m":  round(float(easting),  3),
            })
            n_ok += 1
        print(f"  {sheet_name}: {n_ok} cọc hợp lệ")
    wb.close()
    return records

# ── SQLite ─────────────────────────────────────────────────────────────────────
CREATE_SQL = """
CREATE TABLE IF NOT EXISTS cdm_toado (
    id          INTEGER PRIMARY KEY AUTOINCREMENT,
    zone        TEXT    NOT NULL,
    point_name  TEXT,
    northing_m  REAL    NOT NULL,
    easting_m   REAL    NOT NULL,
    source      TEXT    DEFAULT 'xlsx_toado_cdm'
)
"""

def update_sqlite(records: list[dict], db: Path):
    con = sqlite3.connect(str(db))
    con.execute(CREATE_SQL)
    con.execute("DELETE FROM cdm_toado WHERE source='xlsx_toado_cdm'")
    con.executemany(
        "INSERT INTO cdm_toado (zone, point_name, northing_m, easting_m, source) "
        "VALUES (:zone, :point_name, :northing_m, :easting_m, 'xlsx_toado_cdm')",
        records,
    )
    con.commit()
    # Thống kê per zone
    for zone, in con.execute("SELECT DISTINCT zone FROM cdm_toado").fetchall():
        cnt = con.execute(
            "SELECT COUNT(*) FROM cdm_toado WHERE zone=?", (zone,)
        ).fetchone()[0]
        print(f"  SQLite cdm_toado [{zone}]: {cnt} rows")
    con.close()

# ── JSON ───────────────────────────────────────────────────────────────────────
def update_json(records: list[dict], path: Path):
    zones = {}
    for r in records:
        zones.setdefault(r["zone"], 0)
        zones[r["zone"]] += 1
    out = {
        "_meta": {
            "source":   XLSX.name,
            "updated":  __import__("datetime").date.today().isoformat(),
            "n_total":  len(records),
            "n_by_zone": zones,
        },
        "points": records,
    }
    path.write_text(
        json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(f"  JSON: {path.name}  ({len(records)} cọc)")

# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"Đọc {XLSX.name} ...")
    recs = parse_xlsx(XLSX)
    print(f"Tổng: {len(recs)} cọc CDM hợp lệ")
    update_sqlite(recs, DB)
    update_json(recs, JSON_OUT)
    print("Xong.")
